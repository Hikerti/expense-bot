import pytest
from io import BytesIO
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from src.db.engine import SessionFactory
from src.db.repo import TransactionRepo
from src.core.enums import TransactionStatus
from src.core.schemas import ParsedExpense
from src.services.export import ExportService
from src.services.transaction import TransactionService


class TestExportService:

    def test_export_empty_list(self):
        export = ExportService()
        result = export.generate_xlsx([])

        assert isinstance(result, BytesIO)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "CF"
        assert ws.cell(row=1, column=1).value == "Категория"
        assert ws.cell(row=1, column=7).value == "Валюта"
        assert ws.cell(row=1, column=10).value == "Сумма в USD"

    def test_export_headers_count(self):
        export = ExportService()
        assert len(export.HEADERS) == 13

    @pytest.mark.asyncio
    async def test_export_with_data(self):
        async with SessionFactory() as session:
            svc = TransactionService(
                transaction_repo=TransactionRepo(session),
                reference_repo=ReferenceRepo(session),
            )

            parsed = ParsedExpense(
                amount=Decimal("100.00"),
                currency="USD",
                description="Export test",
                category_code="it_services",
                account_code="falcon_card",
                tranche_code="2025_q1",
                date=date.today(),
                is_recurring=False,
            )
            await svc.create_from_parsed(parsed=parsed)

        async with SessionFactory() as session:
            repo = TransactionRepo(session)
            transactions = await repo.list_for_export()

        export = ExportService()
        result = export.generate_xlsx(transactions)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=1).value is not None